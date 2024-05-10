import pymysql
from openpyxl import load_workbook

from phoneRankDataParse import mysqlConfig


# 解析公司文档
# 读取 Excel 文件
def read_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    return data


# 从数据库中查询 codeName 对应的 id
def get_phone_info_id(cursor, dev_model, flag):
    query_sql = f"SELECT id,cpu,brand,model,codeName,released,cpu_clock,gpu_clock,gpu,resolution,display_diagonal,ram,dimensions,codeName_trim,model_trim FROM phone_info WHERE codeName_trim LIKE %s"
    if flag == 1:
        query_sql = f"SELECT id,cpu,brand,model,codeName,released,cpu_clock,gpu_clock,gpu,resolution,display_diagonal,ram,dimensions,codeName_trim,model_trim FROM phone_info WHERE model_trim LIKE %s"
    cursor.execute(query_sql, ('%' + str(dev_model) + '%',))
    result = cursor.fetchall()
    if result:
        # print(result[0][1])
        return ','.join(str(row[0]) for row in result), result[0][0], result[0][1], result[0][2], result[0][3], \
            result[0][4], result[0][5], result[0][6], result[0][7], result[0][8], result[0][9], result[0][10], \
            result[0][
                11], result[0][12], result[0][13], result[0][14]
    else:
        return 0, 0, "", "", "", "", "", "", "", "", "", "", "", "", "", ""


def get_phone_info_id_only(cursor, dev_model, flag):
    # ios
    query_sql = f"SELECT id,cpu,brand,model,codeName,released,cpu_clock,gpu_clock,gpu,resolution,display_diagonal,ram,dimensions,codeName_trim,model_trim FROM phone_info WHERE codeName_trim = %s"
    if flag == 1:  # 安卓
        query_sql = f"SELECT id,cpu,brand,model,codeName,released,cpu_clock,gpu_clock,gpu,resolution,display_diagonal,ram,dimensions,codeName_trim,model_trim FROM phone_info WHERE model_trim = %s"
    cursor.execute(query_sql, (str(dev_model)))
    result = cursor.fetchall()
    if result:
        return ','.join(str(row[0]) for row in result), result[0][0], result[0][1], result[0][2], result[0][3], \
            result[0][4], result[0][5], result[0][6], result[0][7], result[0][8], result[0][9], result[0][10], \
            result[0][
                11], result[0][12], result[0][13], result[0][14]
    else:
        return 0, 0, "", "", "", "", "", "", "", "", "", "", "", "", "", ""


# 写入数据到 MySQL 数据库
def write_to_mysql(data, host, user, password, database, table_name, flag):
    # 连接到 MySQL 数据库
    conn = pymysql.connect(host=host, user=user, password=password, database=database)
    # 创建一个游标对象
    cursor = conn.cursor()
    # mobile_phones

    cursor.execute("SHOW TABLES LIKE %s", table_name)
    table_exists = cursor.fetchone()
    # 如果存在，则先删除该表
    if table_exists:
        cursor.execute("DROP TABLE " + table_name)

    # 创建表的 SQL 语句
    create_table_sql = f"""CREATE TABLE IF NOT EXISTS {table_name} (
                             id INT AUTO_INCREMENT PRIMARY KEY,
                             dev_model VARCHAR(255),
                             cnt INT,
                             phone_info_ids VARCHAR(255),
                             phone_info_id INT,
                             cpu VARCHAR(255),
                             rank INT,
                             brand VARCHAR(255),
                             model VARCHAR(255),
                             codeName VARCHAR(255),
                             released VARCHAR(255),
                             cpu_clock VARCHAR(255),
                             gpu_clock VARCHAR(255),
                             gpu VARCHAR(255),
                             resolution VARCHAR(255),
                             display_diagonal VARCHAR(255),
                             ram VARCHAR(255),
                             dimensions VARCHAR(255),
                             codeName_trim VARCHAR(255),
                             model_trim VARCHAR(255)
                         )"""
    cursor.execute(create_table_sql)

    # 插入数据的 SQL 语句
    insert_sql = f"""INSERT INTO {table_name} (dev_model, cnt, phone_info_ids,phone_info_id,cpu,brand,rank,model,codeName,released,cpu_clock,gpu_clock,gpu,resolution,display_diagonal,ram,dimensions,codeName_trim,model_trim) VALUES (%s, %s, %s,%s,%s, %s, %s,%s,%s, %s, %s,%s,%s ,%s, %s,%s,%s,%s,%s)"""
    num = 0
    # 将数据逐行插入到数据库中
    for row in data:
        if len(str(row[1])) > 0:
            num += 1
            if num >= 700:  # 安卓只保存top700+数据
                break
            dev_model, cnt = str(row[0]).strip(), str(row[1]).strip()
            phone_info_ids, phone_info_id, cpu, brand, model, codeName, released, cpu_clock, gpu_clock, gpu, resolution, display_diagonal, ram, dimensions, codeName_trim, model_trim = get_phone_info_id_only(
                cursor, dev_model, flag)
            if phone_info_ids == 0:
                phone_info_ids, phone_info_id, cpu, brand, model, codeName, released, cpu_clock, gpu_clock, gpu, resolution, display_diagonal, ram, dimensions, codeName_trim, model_trim = get_phone_info_id(
                    cursor, dev_model, flag)
            cursor.execute(insert_sql, (
                dev_model, cnt, phone_info_ids, phone_info_id, cpu, brand, 0, model, codeName, released, cpu_clock,
                gpu_clock,
                gpu, resolution, display_diagonal, ram, dimensions, codeName_trim, model_trim))

    # 提交事务
    conn.commit()

    # 关闭游标和连接
    cursor.close()
    conn.close()


def main():
    mysqlData = mysqlConfig.get_mysql_config()
    if len(mysqlData) < 3:
        return
    # MySQL 配置信息
    host = mysqlData[0]
    user = mysqlData[1]
    password = mysqlData[2]
    database = 'phone'
    # 1、处理安卓
    # 读取 Excel 文件
    data = read_excel('android.xlsx')
    # 写入数据到 MySQL 数据库
    write_to_mysql(data, host, user, password, database, 'cs_android', 1)
    # 2、处理ios
    data1 = read_excel('ios.xlsx')
    # 写入数据到 MySQL 数据库
    write_to_mysql(data1, host, user, password, database, 'cs_ios', 2)


if __name__ == "__main__":
    main()
